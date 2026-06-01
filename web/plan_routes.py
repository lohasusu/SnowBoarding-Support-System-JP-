"""
plan_routes.py — 整合查詢路由（/plan 頁面 + /api/plan/download）
在 main.py 中以 app.include_router(plan_router) 掛載。
"""
import io
import re
from dataclasses import asdict
from pathlib import Path

from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

BASE_URL = "https://snowboarding-support-system-jp-production.up.railway.app"

plan_router = APIRouter()
TEMPLATE_DIR = Path(__file__).parent / "templates"
_templates   = Jinja2Templates(directory=str(TEMPLATE_DIR))


def _ctx(request: Request, **kw) -> dict:
    return {
        "base_url": BASE_URL,
        "canonical_url": BASE_URL + str(request.url.path),
        **kw,
    }


@plan_router.get("/plan", response_class=HTMLResponse)
async def plan_page(request: Request):
    return _templates.TemplateResponse(request, "plan.html", _ctx(request))


def _airline_name(flights_str: str) -> str:
    legs = re.split(r'\s*→\s*', flights_str or "")
    names = set()
    for leg in legs:
        m = re.match(r'^([A-Za-z][A-Za-z\s]*?)(?=\s*\d|\s*\()', leg.strip())
        if m:
            names.add(m.group(1).strip())
    return " / ".join(sorted(names)) or (flights_str[:25] if flights_str else "")


def _generate_plan_excel(flights: list[dict], ski: list[dict], meta: dict) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: 行程摘要 ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "行程摘要"
    FILL_H = PatternFill("solid", fgColor="1F4E79")
    FONT_H = Font(bold=True, color="FFFFFF")
    summary = [
        ("出發機場", meta.get("origin", "")),
        ("目的地機場", meta.get("destination", "")),
        ("雪場地區", meta.get("region", "") or "全部"),
        ("出發日期", meta.get("departure", "")),
        ("回程日期", meta.get("ret_date", "") or "單程"),
        ("人數", meta.get("adults", 1)),
        ("機票結果數", len(flights)),
        ("雪票結果數", len(ski)),
    ]
    for row_i, (k, v) in enumerate(summary, 1):
        c1 = ws1.cell(row=row_i, column=1, value=k)
        c1.fill = FILL_H; c1.font = FONT_H
        ws1.cell(row=row_i, column=2, value=v)
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 24

    # ── Sheet 2: 航班 ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("機票")
    flight_headers = ["航空公司", "出發", "抵達", "飛行時間", "轉機次數", "票價(TWD)"]
    for col, h in enumerate(flight_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = FILL_H; c.font = FONT_H
    for row_i, r in enumerate(flights, 2):
        stops = r.get("stops", 0)
        ws2.cell(row=row_i, column=1, value=_airline_name(r.get("flights_str", "")))
        ws2.cell(row=row_i, column=2, value=r.get("dep_time", ""))
        ws2.cell(row=row_i, column=3, value=r.get("arr_time", ""))
        ws2.cell(row=row_i, column=4, value=r.get("duration", ""))
        ws2.cell(row=row_i, column=5, value="直飛" if stops == 0 else f"{stops}轉")
        ws2.cell(row=row_i, column=6, value=float(r.get("price") or 0))
    for i, w in enumerate([20, 18, 18, 12, 10, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 雪票 ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("雪票")
    ski_headers = ["雪場", "地區", "票種", "票價", "雪季", "官網"]
    for col, h in enumerate(ski_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = FILL_H; c.font = FONT_H
    for row_i, r in enumerate(ski, 2):
        ws3.cell(row=row_i, column=1, value=r.get("resort", ""))
        ws3.cell(row=row_i, column=2, value=r.get("region", ""))
        ws3.cell(row=row_i, column=3, value=r.get("ticket_type", ""))
        ws3.cell(row=row_i, column=4, value=r.get("price", ""))
        ws3.cell(row=row_i, column=5, value=r.get("season", ""))
        ws3.cell(row=row_i, column=6, value=r.get("source_url", ""))
    for i, w in enumerate([20, 10, 20, 12, 10, 30], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@plan_router.post("/api/plan/download")
async def api_plan_download(request: Request):
    body    = await request.json()
    flights = body.get("flights", [])
    ski     = body.get("ski", [])
    meta    = body.get("meta", {})
    buf     = _generate_plan_excel(flights, ski, meta)
    dep  = meta.get("departure", "unknown")
    orig = meta.get("origin", "")
    dest = meta.get("destination", "")
    fname = f"snowtrip_{orig}-{dest}_{dep}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
