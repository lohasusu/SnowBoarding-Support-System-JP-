"""
SnowTrip Japan — FastAPI Web Application
執行: python main.py  或  uvicorn main:app --reload
"""
import io
import re
import sys
from dataclasses import asdict
from pathlib import Path

# 本地開發：web/ 在 D:\SideProject\web\，ROOT = D:\SideProject\
# 部署環境：web/ 在 snowboarding_support\web\，ROOT = snowboarding_support\
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "flight_search"))

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import uvicorn

BASE_URL = "https://snowboarding-support-system-jp-production.up.railway.app"

app = FastAPI(title="SnowTrip Japan", docs_url=None, redoc_url=None)

TEMPLATE_DIR = Path(__file__).parent / "templates"
STATIC_DIR   = Path(__file__).parent / "static"

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
templates = Jinja2Templates(directory=str(TEMPLATE_DIR))


def _ctx(request: Request, **kw) -> dict:
    return {
        "base_url": BASE_URL,
        "canonical_url": BASE_URL + str(request.url.path),
        **kw,
    }


# ── 頁面路由 ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(request, "index.html", _ctx(request))


@app.get("/ski", response_class=HTMLResponse)
async def ski_page(request: Request):
    return templates.TemplateResponse(request, "ski.html", _ctx(request))


@app.get("/flight", response_class=HTMLResponse)
async def flight_page(request: Request):
    return templates.TemplateResponse(request, "flight.html", _ctx(request))


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse(request, "auth/login.html", _ctx(request))


@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    return templates.TemplateResponse(request, "auth/register.html", _ctx(request))


# ── API：雪票 ─────────────────────────────────────────────────────────────────

def _import_ski():
    try:
        from snowboarding_support.scraper import get_ticket_prices
    except ImportError:
        from scraper import get_ticket_prices  # 部署環境：ROOT 即 snowboarding_support/
    return get_ticket_prices


@app.get("/api/ski/search")
async def api_ski_search(region: str = None, name: str = None):
    try:
        get_ticket_prices = _import_ski()
        results = get_ticket_prices(region=region or None, name=name or None)
        return {"ok": True, "data": [asdict(r) for r in results]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/ski/download")
async def api_ski_download(region: str = None, name: str = None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        get_ticket_prices = _import_ski()

        results = get_ticket_prices(region=region or None, name=name or None)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "雪票價格"

        headers = ["雪場", "地區", "票種（日文）", "票種（中文）", "票價", "雪季", "查詢時間", "票價頁連結"]
        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_i, r in enumerate(results, 2):
            d = asdict(r)
            ws.cell(row=row_i, column=1, value=d.get("resort", ""))
            ws.cell(row=row_i, column=2, value=d.get("region", ""))
            ws.cell(row=row_i, column=3, value=d.get("ticket_type", ""))
            ws.cell(row=row_i, column=4, value=d.get("ticket_type_zh", ""))
            ws.cell(row=row_i, column=5, value=d.get("price", ""))
            ws.cell(row=row_i, column=6, value=d.get("season", ""))
            ws.cell(row=row_i, column=7, value=d.get("scraped_at", ""))
            ws.cell(row=row_i, column=8, value=d.get("source_url", ""))

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"ski_prices_{region or 'all'}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        return Response(content=str(e), status_code=500)


# ── API：機票 ─────────────────────────────────────────────────────────────────

@app.get("/api/flight/search")
async def api_flight_search(
    origin: str = "TPE",
    destination: str = "CTS",
    dest_name: str = "新千歲",
    departure: str = None,
    ret_date: str = None,
    currency: str = "TWD",
    adults: int = 1,
):
    if not departure:
        return {"ok": False, "error": "請輸入出發日期"}
    try:
        import os
        from dotenv import load_dotenv
        # 載入 flight_search/.env
        env_path = ROOT / "flight_search" / ".env"
        load_dotenv(dotenv_path=env_path)

        serpapi_key = os.getenv("SERPAPI_API_KEY", "")
        backend = None
        backend_name = "unknown"

        if serpapi_key:
            from backends.serpapi_backend import SerpApiBackend
            _b = SerpApiBackend(api_key=serpapi_key)
            if _b.is_available():
                backend = _b
                backend_name = "SerpAPI"

        if not backend:
            from backends.fast_flights_backend import FastFlightsBackend
            backend = FastFlightsBackend()
            backend_name = "fast-flights (fallback)"

        results = backend.search(
            origin=origin,
            destination=destination,
            dest_name=dest_name,
            departure_date=departure,
            return_date=ret_date or None,
            currency="TWD",
            adults=adults,
        )
        return {"ok": True, "backend": backend_name, "data": [asdict(r) for r in results]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── 機票 Excel 下載 ───────────────────────────────────────────────────────────

def _airline_name(flights_str: str) -> str:
    """從 flights_str 擷取航空公司名稱，去除航班號與路線"""
    legs = re.split(r'\s*→\s*', flights_str or "")
    names = set()
    for leg in legs:
        m = re.match(r'^([A-Za-z][A-Za-z\s]*?)(?=\s*\d|\s*\()', leg.strip())
        if m:
            names.add(m.group(1).strip())
    return " / ".join(sorted(names)) or flights_str[:25]


def _generate_flight_excel(flights: list[dict], meta: dict) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    is_rt = bool(meta.get("ret_date"))

    # ── 顏色定義 ──────────────────────────────────────────────────────────────
    FILL_BANNER   = PatternFill("solid", fgColor="D6E4F0")
    FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")
    FILL_ROW_TOP  = PatternFill("solid", fgColor="C6EFCE")  # 前3名
    FILL_ROW_REST = PatternFill("solid", fgColor="E2EFDA")  # 其餘
    FILL_DEP_PRICE= PatternFill("solid", fgColor="BDD7EE")  # 去程票價欄
    FILL_RET_PRICE= PatternFill("solid", fgColor="F9CBAD")  # 回程票價欄
    FILL_TOTAL_1  = PatternFill("solid", fgColor="70AD47")  # 最低合計
    FILL_TOTAL_N  = PatternFill("solid", fgColor="A9D18E")  # 其他合計

    FONT_BANNER = Font(bold=True, size=10, color="1F4E79")
    FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
    FONT_DATA   = Font(size=12)
    FONT_TOTAL1 = Font(bold=True, size=13, color="375623")
    FONT_TOTALN = Font(bold=True, size=13, color="375623")
    FONT_NOTE   = Font(size=9, color="595959")

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "航班搜尋結果"

    # ── Row 1：搜尋摘要 ────────────────────────────────────────────────────────
    origin   = meta.get("origin", "")
    dest     = meta.get("destination", "")
    dep      = meta.get("departure", "")
    ret      = meta.get("ret_date", "")
    adults   = meta.get("adults", 1)
    banner   = (f"✈  {origin} → {dest}   去程: {dep}"
                + (f"   回程: {ret}" if is_rt else "")
                + f"   乘客: {adults} 人   資料來源: Google Flights")
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = banner
    c.fill  = FILL_BANNER
    c.font  = FONT_BANNER
    c.alignment = CENTER

    # ── Row 2：免責聲明 ────────────────────────────────────────────────────────
    ws.merge_cells("A2:M2")
    c = ws["A2"]
    c.value = "組合推薦（依合計票價排序）  |  資料來源: Google Flights  |  票價僅供參考，請至官網確認"
    c.fill  = FILL_BANNER
    c.font  = FONT_BANNER
    c.alignment = CENTER

    # ── Row 3：空白 ────────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── Row 4：欄位標題 ───────────────────────────────────────────────────────
    HEADERS = ["排名","航空公司","去程出發","去程抵達","去程時長","去程轉機",
               "去程票價(TWD)","回程出發","回程抵達","回程時長","回程轉機",
               "回程票價(TWD)","合計票價(TWD)"]
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = CENTER

    # ── 依合計票價排序 ─────────────────────────────────────────────────────────
    sorted_flights = sorted(flights, key=lambda r: float(r.get("price") or 0))

    for rank, r in enumerate(sorted_flights, 1):
        row = rank + 4
        row_fill = FILL_ROW_TOP if rank <= 3 else FILL_ROW_REST
        stops_str = "直飛" if (r.get("stops") or 0) == 0 else f"{r.get('stops')}轉"
        ret_stops = r.get("ret_stops")
        ret_stops_str = ("直飛" if ret_stops == 0 else f"{ret_stops}轉") if ret_stops is not None else "—"

        vals = [
            rank,
            _airline_name(r.get("flights_str", "")),
            r.get("dep_time") or "—",
            r.get("arr_time") or "—",
            r.get("duration") or "—",
            stops_str,
            float(r.get("price") or 0) if not is_rt else "—",
            r.get("ret_dep_time") or "—" if is_rt else "—",
            r.get("ret_arr_time") or "—" if is_rt else "—",
            r.get("ret_duration") or "—" if is_rt else "—",
            ret_stops_str if is_rt else "—",
            "—",
            float(r.get("price") or 0),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = FONT_DATA
            c.alignment = CENTER
            # 去程票價欄（G=7）
            if col == 7:
                c.fill = FILL_DEP_PRICE
            # 回程票價欄（L=12）
            elif col == 12:
                c.fill = FILL_RET_PRICE
            # 合計票價欄（M=13）
            elif col == 13:
                c.fill = FILL_TOTAL_1 if rank == 1 else FILL_TOTAL_N
                c.font = FONT_TOTAL1 if rank == 1 else FONT_TOTALN
            else:
                c.fill = row_fill

    # ── 頁尾說明 ──────────────────────────────────────────────────────────────
    footer_row = len(sorted_flights) + 6
    ws.cell(row=footer_row, column=2, value="合計最低前3名").font = FONT_NOTE
    ws.cell(row=footer_row, column=4, value="其餘組合").font = FONT_NOTE

    # ── 欄寬 ──────────────────────────────────────────────────────────────────
    col_widths = [6, 20, 18, 18, 12, 10, 16, 18, 18, 12, 10, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.post("/api/flight/download")
async def api_flight_download(request: Request):
    body    = await request.json()
    flights = body.get("flights", [])
    meta    = body.get("meta", {})
    buf     = _generate_flight_excel(flights, meta)
    dep  = meta.get("departure", "unknown")
    orig = meta.get("origin", "")
    dest = meta.get("destination", "")
    fname = f"flights_{orig}-{dest}_{dep}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── 診斷 ─────────────────────────────────────────────────────────────────────

@app.get("/api/env-check")
async def env_check():
    import os
    from dotenv import load_dotenv
    env_path = ROOT / "flight_search" / ".env"
    load_dotenv(dotenv_path=env_path)

    key = os.getenv("SERPAPI_API_KEY", "")

    # 找所有含 SERP / API / KEY 關鍵字的環境變數（不顯示值，只顯示名稱）
    related_keys = [
        k for k in os.environ
        if any(w in k.upper() for w in ("SERP", "FLIGHT", "GOOGLE"))
    ]

    return {
        "serpapi_key_set": bool(key),
        "serpapi_key_length": len(key),
        "serpapi_key_prefix": key[:6] + "..." if key else "(empty)",
        "env_file_exists": env_path.exists(),
        "ROOT": str(ROOT),
        "related_env_keys_found": sorted(related_keys),
        "total_env_vars": len(os.environ),
    }


# ── SEO ───────────────────────────────────────────────────────────────────────

@app.get("/robots.txt", response_class=Response)
async def robots():
    return Response(
        content=f"User-agent: *\nAllow: /\nDisallow: /api/\n\nSitemap: {BASE_URL}/sitemap.xml\n",
        media_type="text/plain",
    )


@app.get("/sitemap.xml", response_class=Response)
async def sitemap():
    pages = [("/", "1.0", "weekly"), ("/ski", "0.9", "daily"), ("/flight", "0.8", "weekly")]
    items = "\n".join(
        f"  <url><loc>{BASE_URL}{p}</loc><changefreq>{cf}</changefreq><priority>{pr}</priority></url>"
        for p, pr, cf in pages
    )
    xml = f'<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n{items}\n</urlset>'
    return Response(content=xml, media_type="application/xml")


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
