"""
ski_early_bird_scraper.py
爬取日本滑雪場官網的票種與票價，輸出至 Excel。

目標網址來源：urls.json（與本腳本同目錄）
只處理已設定 ticket_url 的雪場；無 ticket_url 者記錄至「略過雪場」sheet。
"""

import argparse
import asyncio
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

# Windows 終端機預設 cp950，強制 utf-8 避免中文/符號 UnicodeEncodeError
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from deep_translator import GoogleTranslator
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

# ── 雪季判斷 ─────────────────────────────────────────────────────────────────

def get_current_season() -> str:
    """
    1–7 月 → 上一個雪季 (去年/今年)，例：2026-05 → "25/26"
    8–12 月 → 當前雪季 (今年/明年)，例：2026-10 → "26/27"
    """
    now = datetime.now()
    y = now.year
    if now.month <= 7:
        return f"{(y - 1) % 100:02d}/{y % 100:02d}"
    else:
        return f"{y % 100:02d}/{(y + 1) % 100:02d}"


# ── 設定區 ──────────────────────────────────────────────────────────────────

OUTPUT_DIR = Path(__file__).parent / "output"
URLS_FILE  = Path(__file__).parent / "urls.json"

_CURRENT_SEASON: str = get_current_season()

# 票種相關關鍵字（日英），用於精準模式過濾
TICKET_KEYWORDS = [
    "券", "パス", "チケット", "リフト", "大人", "子供", "こども", "シニア",
    "1日", "半日", "通し", "シーズン", "早割", "前売", "当日", "ナイター",
    "pass", "ticket", "adult", "child", "senior", "day", "half", "season",
    "lift", "early bird", "discount",
]

# 日圓價格正則
PRICE_PATTERN = re.compile(
    r"[¥￥][\d,]+|[\d,]+\s*円|JPY\s*[\d,]+|[\d,]+\s*JPY",
    re.IGNORECASE,
)


# ── 資料結構 ─────────────────────────────────────────────────────────────────

@dataclass
class TicketItem:
    resort: str
    region: str
    source_url: str
    ticket_type: str
    ticket_type_zh: str
    price: str
    season: str = field(default_factory=lambda: _CURRENT_SEASON)
    scraped_at: str = field(
        default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )


# ── 翻譯工具 ─────────────────────────────────────────────────────────────────

_translate_cache: dict[str, str] = {}
_translator = GoogleTranslator(source="ja", target="zh-TW")


def translate_ja_to_zh(text: str) -> str:
    """日文 → 繁體中文，失敗時回傳原文。結果快取避免重複呼叫。"""
    if text in _translate_cache:
        return _translate_cache[text]
    try:
        result = _translator.translate(text) or text
    except Exception:
        result = text
    _translate_cache[text] = result
    return result


# ── 工具函式 ─────────────────────────────────────────────────────────────────

def extract_prices(text: str) -> list[str]:
    return [m.group().strip() for m in PRICE_PATTERN.finditer(text)]


def is_ticket_related(text: str) -> bool:
    t = text.lower()
    return any(kw.lower() in t for kw in TICKET_KEYWORDS)


def load_targets(region: str | None = None, name: str | None = None) -> list[dict]:
    """urls.json を読む。region / name で絞り込み可。"""
    if not URLS_FILE.exists():
        print(f"[警告] 找不到 {URLS_FILE.name}，無法執行。")
        return []

    try:
        data = json.loads(URLS_FILE.read_text(encoding="utf-8"))
        targets = [t for t in data if isinstance(t, dict) and "name" in t and "url" in t]
    except (json.JSONDecodeError, OSError) as e:
        print(f"[警告] 讀取 {URLS_FILE.name} 失敗（{e}）。")
        return []

    if not targets:
        print(f"[警告] {URLS_FILE.name} 無有效目標。")
        return []

    if region:
        targets = [t for t in targets if region in t.get("note", "")]
    if name:
        name_lower = name.lower()
        targets = [t for t in targets if name_lower in t["name"].lower()]

    suffix = ""
    if region:
        suffix += f"，地區篩選：{region}"
    if name:
        suffix += f"，名稱篩選：{name}"
    print(f"[設定] 從 {URLS_FILE.name} 載入 {len(targets)} 個目標{suffix}")
    return targets


# ── 精準選擇器擷取 ────────────────────────────────────────────────────────────

async def _extract_by_selectors(page, resort, region, url, seen, selectors: dict) -> list[TicketItem]:
    """
    依 urls.json 中手動設定的 selectors 精準擷取。
    支援三種結構：table、dl、card。
    """
    items = []
    sel_type = selectors.get("type", "table")
    container_sel = selectors.get("container", "")

    if not container_sel:
        return items

    try:
        containers = await page.query_selector_all(container_sel)
    except Exception:
        return items

    for container in containers:
        if sel_type == "dl":
            dts = await container.query_selector_all("dt")
            for dt in dts:
                try:
                    dt_text = (await dt.inner_text()).strip()
                    dd_el   = await dt.evaluate_handle("el => el.nextElementSibling")
                    dd_text = (await dd_el.inner_text()).strip() if dd_el else ""
                except Exception:
                    continue
                prices = extract_prices(dd_text) or extract_prices(dt_text)
                if prices and is_ticket_related(dt_text):
                    for price in prices:
                        key = f"{resort}||{dt_text}||{price}"
                        if key not in seen:
                            seen.add(key)
                            items.append(TicketItem(
                                resort, region, url,
                                dt_text, translate_ja_to_zh(dt_text), price,
                            ))

        elif sel_type == "card":
            name_sel  = selectors.get("ticket_name", "")
            price_sel = selectors.get("ticket_price", "")
            if not name_sel or not price_sel:
                continue
            try:
                name_el  = await container.query_selector(name_sel)
                price_el = await container.query_selector(price_sel)
                if not name_el or not price_el:
                    continue
                ticket_type = (await name_el.inner_text()).strip()
                price_text  = (await price_el.inner_text()).strip()
            except Exception:
                continue
            prices = extract_prices(price_text)
            if prices and is_ticket_related(ticket_type):
                for price in prices:
                    key = f"{resort}||{ticket_type}||{price}"
                    if key not in seen:
                        seen.add(key)
                        items.append(TicketItem(
                            resort, region, url,
                            ticket_type, translate_ja_to_zh(ticket_type), price,
                        ))

        else:  # table
            row_sel = selectors.get("row", "tr")
            try:
                rows = await container.query_selector_all(row_sel)
            except Exception:
                continue
            for row in rows:
                cells = await row.query_selector_all("td, th")
                cell_texts = []
                for c in cells:
                    try:
                        cell_texts.append((await c.inner_text()).strip())
                    except Exception:
                        cell_texts.append("")
                row_text = " ".join(cell_texts)
                prices = extract_prices(row_text)
                if not prices:
                    continue
                ticket_type = next(
                    (t for t in cell_texts if t and not extract_prices(t)), ""
                )
                if not ticket_type or not is_ticket_related(ticket_type):
                    continue
                for price in prices:
                    key = f"{resort}||{ticket_type}||{price}"
                    if key not in seen:
                        seen.add(key)
                        items.append(TicketItem(
                            resort, region, url,
                            ticket_type, translate_ja_to_zh(ticket_type), price,
                        ))

    return items


async def extract_tickets(page, url: str, resort: str, region: str, seen: set,
                          selectors: dict | None = None) -> list[TicketItem]:
    """有 selectors → 精準模式；無 selectors → 跳過（不使用通用掃描）。"""
    if not selectors:
        return []
    return await _extract_by_selectors(page, resort, region, url, seen, selectors)


# ── 爬蟲核心 ─────────────────────────────────────────────────────────────────

async def scrape_resort(browser, target: dict) -> list[TicketItem]:
    """爬取單一雪場票價。必須設定 ticket_url + selectors 才會產生資料。"""
    resort     = target["name"]
    region     = target.get("note", "")
    ticket_url = target.get("ticket_url") or None
    selectors  = target.get("selectors") or None

    if not ticket_url:
        print(f"[{resort}] 尚未設定 ticket_url，略過")
        return []

    if selectors:
        selectors = {k: v for k, v in selectors.items() if not k.startswith("_")}

    mode_tag = "[精準]" if selectors else "[無 selector]"
    seen: set[str] = set()
    page = await browser.new_page()

    try:
        print(f"[{resort}]{mode_tag} 開啟: {ticket_url}")
        await page.goto(ticket_url, timeout=45_000, wait_until="domcontentloaded")
        await page.wait_for_selector("body", timeout=10_000)

        results = await extract_tickets(page, ticket_url, resort, region, seen, selectors)

        if results:
            print(f"[{resort}] 擷取 {len(results)} 筆")
        else:
            print(f"[{resort}] 未找到票價資料")

        return results

    except PlaywrightTimeoutError:
        print(f"[{resort}] 頁面載入逾時。")
        return []
    except Exception as e:
        print(f"[{resort}] 發生錯誤: {e}")
        return []
    finally:
        await page.close()


# ── Excel 格式模板 ────────────────────────────────────────────────────────────

# 地區色碼（與 REGION_ORDER 對應）
REGION_ORDER = ["北海道", "長野", "新潟", "山形", "青森", "福島"]
REGION_COLORS = {
    "北海道": "BDD7EE",  # 藍
    "長野":   "C6E0B4",  # 綠
    "新潟":   "FFE699",  # 黃
    "山形":   "D9B1F0",  # 紫
    "青森":   "F4B183",  # 橘
    "福島":   "FF9999",  # 紅
}

# 早鳥判斷關鍵字（日英）
EARLY_BIRD_KEYWORDS = [
    "早割", "早鳥", "前売", "前売り", "早期",
    "early bird", "advance", "pre-sale", "pre sale",
]

TICKET_COLS = ["地區", "雪場名稱", "票種分類", "票種（日文）", "票種（中文）", "票價", "雪季", "票價頁面", "抓取時間"]
SKIP_COLS   = ["地區", "雪場名稱", "首頁 URL", "原因"]

COL_WIDTHS = {
    "地區":       14,
    "雪場名稱":   24,
    "票種分類":   10,
    "票種（日文）": 32,
    "票種（中文）": 32,
    "票價":       14,
    "雪季":        8,
    "票價頁面":   55,
    "抓取時間":   20,
    "首頁 URL":   55,
    "原因":       30,
}

_HEADER_FILL     = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT     = Font(bold=True, color="FFFFFF", size=11)
_EARLY_FILL      = PatternFill("solid", fgColor="E2EFDA")  # 淺綠：早鳥
_ONSITE_FILL     = PatternFill("solid", fgColor="FFF2CC")  # 淺黃：現場
_HYPERLINK_FONT  = Font(color="0563C1", underline="single")
_BOLD_FONT       = Font(bold=True)
_CENTER          = Alignment(horizontal="center", vertical="center")
_RIGHT           = Alignment(horizontal="right",  vertical="center")


def classify_ticket_type(ticket_type: str) -> str:
    """票種分類：早鳥 / 現場"""
    t = ticket_type.lower()
    if any(kw.lower() in t for kw in EARLY_BIRD_KEYWORDS):
        return "早鳥"
    return "現場"


def _region_sort_key(region: str) -> int:
    for i, r in enumerate(REGION_ORDER):
        if r in region:
            return i
    return 99


def _region_fill(region: str) -> PatternFill:
    for key, color in REGION_COLORS.items():
        if key in region:
            return PatternFill("solid", fgColor=color)
    return PatternFill("solid", fgColor="EEEEEE")


def _write_header(ws, cols: list[str]) -> None:
    for col, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        ws.column_dimensions[c.column_letter].width = COL_WIDTHS.get(h, 18)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def save_to_excel(
    items: list[TicketItem],
    output_dir: Path,
    skipped: list[dict] | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    seq = 1
    while True:
        filepath = output_dir / f"SNOWBOARD_TICKETS_{today}_{seq:03d}.xlsx"
        if not filepath.exists():
            break
        seq += 1

    wb = openpyxl.Workbook()

    # ── Sheet 1：票種票價 ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "票種票價"
    _write_header(ws, TICKET_COLS)

    # 排序：地區順序 → 雪場名稱 → 早鳥優先 → 票種原文
    sorted_items = sorted(items, key=lambda x: (
        _region_sort_key(x.region),
        x.resort,
        0 if classify_ticket_type(x.ticket_type) == "早鳥" else 1,
        x.ticket_type,
    ))

    prev_resort = None
    for row_idx, item in enumerate(sorted_items, 2):
        category  = classify_ticket_type(item.ticket_type)
        row_fill  = _EARLY_FILL if category == "早鳥" else _ONSITE_FILL
        rfill     = _region_fill(item.region)

        # 地區（色碼）
        c = ws.cell(row=row_idx, column=1, value=item.region)
        c.fill = rfill
        c.alignment = _CENTER

        # 雪場名稱（雪場第一筆加粗）
        c = ws.cell(row=row_idx, column=2, value=item.resort)
        if item.resort != prev_resort:
            c.font = _BOLD_FONT
            prev_resort = item.resort

        # 票種分類（色碼）
        c = ws.cell(row=row_idx, column=3, value=category)
        c.fill = row_fill
        c.alignment = _CENTER

        # 票種（日文）
        ws.cell(row=row_idx, column=4, value=item.ticket_type)

        # 票種（中文）
        ws.cell(row=row_idx, column=5, value=item.ticket_type_zh)

        # 票價
        c = ws.cell(row=row_idx, column=6, value=item.price)
        c.alignment = _RIGHT

        # 雪季
        c = ws.cell(row=row_idx, column=7, value=item.season)
        c.alignment = _CENTER

        # 票價頁面（可點擊超連結）
        c = ws.cell(row=row_idx, column=8, value=item.source_url)
        c.hyperlink = item.source_url
        c.font = _HYPERLINK_FONT

        # 抓取時間
        c = ws.cell(row=row_idx, column=9, value=item.scraped_at)
        c.alignment = _CENTER

    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2：略過雪場 ────────────────────────────────────────────────────
    if skipped:
        ws2 = wb.create_sheet("略過雪場")
        _write_header(ws2, SKIP_COLS)

        skip_sorted = sorted(skipped, key=lambda t: (
            _region_sort_key(t.get("note", "")),
            t.get("name", ""),
        ))

        for row_idx, t in enumerate(skip_sorted, 2):
            rfill = _region_fill(t.get("note", ""))

            c = ws2.cell(row=row_idx, column=1, value=t.get("note", ""))
            c.fill = rfill
            c.alignment = _CENTER

            ws2.cell(row=row_idx, column=2, value=t.get("name", ""))

            url = t.get("url", "")
            c = ws2.cell(row=row_idx, column=3, value=url)
            if url:
                c.hyperlink = url
                c.font = _HYPERLINK_FONT

            ws2.cell(row=row_idx, column=4, value=t.get("_reason", ""))

        ws2.auto_filter.ref = ws2.dimensions

    wb.save(filepath)
    return filepath


# ── 資料庫預留介面 ────────────────────────────────────────────────────────────

def save_to_database(items: list[TicketItem]) -> None:
    """預留的資料庫寫入函式，補上連線邏輯即可啟用。"""
    if not items:
        return
    print(f"[DB] 預留介面：共 {len(items)} 筆待寫入資料庫（尚未實作）。")


# ── 公開 API（供其他模組 import） ─────────────────────────────────────────────

async def get_ticket_prices_async(region: str | None = None, name: str | None = None) -> list[TicketItem]:
    global _CURRENT_SEASON
    _CURRENT_SEASON = get_current_season()

    targets = load_targets(region=region, name=name)
    active  = [t for t in targets if t.get("ticket_url")]

    all_items: list[TicketItem] = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            tasks = [scrape_resort(browser, t) for t in active]
            for result in await asyncio.gather(*tasks):
                all_items.extend(result)
        finally:
            await browser.close()

    return all_items


def get_ticket_prices(region: str | None = None, name: str | None = None) -> list[TicketItem]:
    """
    公開 API：爬取票價並回傳 TicketItem 列表。
    供其他模組 import 使用，不輸出 Excel。

    範例：
        from snowboarding_support.scraper import get_ticket_prices
        prices = get_ticket_prices(region="北海道")
    """
    return asyncio.run(get_ticket_prices_async(region, name))


# ── CLI 參數 ──────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="日本滑雪場票價爬蟲",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--region", "-r",
        metavar="地區",
        help="地區篩選（北海道 / 長野 / 新潟 / 山形 / 青森 / 福島）\n例：-r 長野",
    )
    parser.add_argument(
        "--name", "-n",
        metavar="名稱",
        help="雪場名稱篩選（部分比對，大小寫不限）\n例：-n furano  或  -n 志賀",
    )
    return parser.parse_args()


# ── 主流程 ────────────────────────────────────────────────────────────────────

async def main() -> None:
    global _CURRENT_SEASON
    args = parse_args()

    _CURRENT_SEASON = get_current_season()
    print(f"[雪季] {_CURRENT_SEASON}")

    targets         = load_targets(region=args.region, name=args.name)
    skipped_targets = [t for t in targets if not t.get("ticket_url")]
    active_targets  = [t for t in targets if t.get("ticket_url")]

    for t in skipped_targets:
        print(f"[{t['name']}] 尚未設定 ticket_url，略過")

    all_items: list[TicketItem] = []
    results: list[list[TicketItem]] = []

    if active_targets:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            try:
                tasks = [scrape_resort(browser, t) for t in active_targets]
                results = list(await asyncio.gather(*tasks))
            finally:
                await browser.close()

        for result in results:
            all_items.extend(result)

    # 有 ticket_url 但抓不到資料的雪場
    no_data_targets = [
        {**t, "_reason": "有 ticket_url 但未抓到票價"}
        for t, result in zip(active_targets, results)
        if not result
    ]

    all_skipped = [
        {**t, "_reason": "尚未設定 ticket_url"} for t in skipped_targets
    ] + no_data_targets

    excel_path = save_to_excel(all_items, OUTPUT_DIR, skipped=all_skipped)
    print(f"\n✓ Excel 已輸出：{excel_path}")
    print(f"  票價資料：{len(all_items)} 筆，略過/未抓到：{len(all_skipped)} 個雪場")

    save_to_database(all_items)


if __name__ == "__main__":
    asyncio.run(main())
