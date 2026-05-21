"""
generate_review_excel.py
讀取 site_analysis/analysis_report.json，
輸出供人工審查的 Excel：所有 URL 均可直接點擊開啟，
並附上選擇器建議和票價樣本。
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ANALYSIS_DIR = Path(__file__).parent / "analysis"
REPORT_FILE  = ANALYSIS_DIR / "analysis_report.json"
OUTPUT_DIR   = Path(__file__).parent / "output"

STATUS_ZH = {
    "analyzed":       "✅ 成功",
    "no_price_found": "⚠️ 未找到票價",
    "timeout":        "❌ 逾時",
    "error":          "❌ 錯誤",
    "pending":        "⏳ 待分析",
}

STATUS_COLOR = {
    "analyzed":       "C6EFCE",   # 綠
    "no_price_found": "FFEB9C",   # 黃
    "timeout":        "FFC7CE",   # 紅
    "error":          "FFC7CE",   # 紅
    "pending":        "DDDDDD",   # 灰
}

HEADERS = [
    "地區", "雪場名稱", "首頁 URL", "票價頁面 URL",
    "狀態", "選擇器類型", "Container 選擇器",
    "票價樣本 1", "票價樣本 2", "錯誤訊息",
]

COL_WIDTHS = [18, 22, 55, 55, 14, 12, 35, 55, 55, 40]

LINK_FONT   = Font(color="0563C1", underline="single")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")


def _set_hyperlink(cell, url: str, display: str | None = None):
    cell.value     = display or url
    cell.hyperlink = url
    cell.font      = LINK_FONT


def _best_samples(hints: list[dict]) -> tuple[str, str]:
    """從 selector_hints 取前兩個不重複、有實質內容的樣本。"""
    seen: set[str] = set()
    samples: list[str] = []
    for h in hints:
        if h.get("error"):
            continue
        for s in h.get("samples", []):
            s = s.strip()
            if s and s not in seen:
                seen.add(s)
                samples.append(s)
            if len(samples) >= 2:
                break
        if len(samples) >= 2:
            break
    s1 = samples[0] if len(samples) > 0 else ""
    s2 = samples[1] if len(samples) > 1 else ""
    return s1, s2


def generate(report_path: Path = REPORT_FILE, output_dir: Path = OUTPUT_DIR) -> Path:
    if not report_path.exists():
        raise FileNotFoundError(f"找不到報告檔案：{report_path}")

    reports: list[dict] = json.loads(report_path.read_text(encoding="utf-8"))
    output_dir.mkdir(parents=True, exist_ok=True)

    today = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = output_dir / f"RESORT_REVIEW_{today}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "網站審查"
    ws.freeze_panes = "A2"

    # 標題列
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    # 資料列
    for row_idx, r in enumerate(reports, 2):
        status   = r.get("status", "pending")
        hints    = r.get("selector_hints") or []
        sel      = r.get("selectors") or {}
        sample1, sample2 = _best_samples(hints)
        hint_types = list({h.get("type", "") for h in hints if not h.get("error") and h.get("type")})

        row_fill = PatternFill("solid", fgColor=STATUS_COLOR.get(status, "FFFFFF"))

        def _cell(col, value=None):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.alignment = WRAP_ALIGN
            c.fill      = row_fill
            return c

        _cell(1, r.get("note", ""))
        _cell(2, r.get("name", ""))

        # 首頁 URL（超連結）
        base_url = r.get("url", "")
        if base_url:
            c = _cell(3)
            _set_hyperlink(c, base_url)
            c.fill = row_fill
        else:
            _cell(3, "")

        # 票價頁面 URL（超連結）
        ticket_url = r.get("ticket_url") or ""
        if ticket_url:
            c = _cell(4)
            _set_hyperlink(c, ticket_url)
            c.fill = row_fill
        else:
            _cell(4, "（未找到，請人工確認）" if status == "analyzed" else "")

        _cell(5, STATUS_ZH.get(status, status))
        _cell(6, ", ".join(hint_types) if hint_types else "")

        container = sel.get("container", "")
        _cell(7, container)
        _cell(8, sample1)
        _cell(9, sample2)
        _cell(10, r.get("error") or "")

        ws.row_dimensions[row_idx].height = 40

    # 統計列（最底部）
    last_row = len(reports) + 3
    total     = len(reports)
    success   = sum(1 for r in reports if r.get("status") == "analyzed")
    no_price  = sum(1 for r in reports if r.get("status") == "no_price_found")
    errors    = sum(1 for r in reports if r.get("status") in ("timeout", "error"))

    ws.cell(row=last_row, column=1, value="統計").font = Font(bold=True)
    ws.cell(row=last_row, column=2,
            value=f"共 {total} 個 ｜ ✅ 成功 {success} ｜ ⚠️ 未找到票價 {no_price} ｜ ❌ 錯誤 {errors}")

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    path = generate()
    print(f"Excel 已輸出：{path}")
