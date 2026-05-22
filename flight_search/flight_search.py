"""
航班搜尋工具
搜尋台灣出發至日本的最便宜航班，支援去回程、預算篩選、Excel 輸出

後端優先順序：
  1. Google Flights (fast-flights)  ← 預設，無需 API Key
  2. SerpAPI (Google Flights)       ← 設定 SERPAPI_API_KEY 後啟用
  3. Travelpayouts                  ← 設定 TRAVELPAYOUTS_TOKEN 後啟用

測試模式：python flight_search.py --mock  （使用假資料，完全不需要 Key）
"""

import argparse
import io
import os
import sys
from datetime import datetime, date

# Windows cp950 → UTF-8
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill, PatternFill, Side
    )
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.utils import get_column_letter
except ImportError:
    print("請先安裝相依套件: pip install -r requirements.txt")
    sys.exit(1)

from backends.base import FlightResult

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ── 機場資料 ──────────────────────────────────────────────────────────────────

ORIGIN_AIRPORTS = {
    "1": ("台北桃園", "TPE"),
    "2": ("台北松山", "TSA"),
    "3": ("高雄小港", "KHH"),
    "4": ("台中清泉崗", "RMQ"),
}

JAPAN_AIRPORTS = {
    "1":  ("東京成田", "NRT"),
    "2":  ("東京羽田", "HND"),
    "3":  ("大阪關西", "KIX"),
    "4":  ("大阪伊丹", "ITM"),
    "5":  ("名古屋中部", "NGO"),
    "6":  ("札幌新千歲", "CTS"),
    "7":  ("福岡", "FUK"),
    "8":  ("沖繩那霸", "OKA"),
    "9":  ("廣島", "HIJ"),
    "10": ("仙台", "SDJ"),
    "11": ("高松", "TAK"),
    "12": ("長崎", "NGS"),
}

CURRENCY_OPTIONS = {
    "1": ("新台幣", "TWD"),
    "2": ("日圓",   "JPY"),
    "3": ("美元",   "USD"),
}


# ── Backend 初始化 ─────────────────────────────────────────────────────────────

def build_backends(mock_mode: bool = False) -> list:
    backends = []

    if mock_mode:
        from backends.mock_backend import MockBackend
        backends.append(MockBackend())
        print("  ✓ Mock 模式（假資料，測試用）")
        return backends

    # Google Flights via fast-flights（無需 Key）
    try:
        from backends.fast_flights_backend import FastFlightsBackend
        b = FastFlightsBackend()
        if b.is_available():
            backends.append(b)
            print(f"  ✓ Google Flights (fast-flights) 已載入  [今日已用 {b.daily_count} 次]")
        else:
            print("  ✗ fast-flights：請先安裝套件 (pip install fast-flights)")
    except Exception as e:
        print(f"  ✗ Google Flights：{e}")

    # SerpAPI（可選，設定 SERPAPI_API_KEY 後啟用）
    serpapi_key = os.environ.get("SERPAPI_API_KEY", "").strip()
    if serpapi_key:
        from backends.serpapi_backend import SerpApiBackend
        b = SerpApiBackend(serpapi_key)
        if b.is_available():
            backends.append(b)
            print("  ✓ SerpAPI (Google Flights) 已載入")
        else:
            print("  ✗ SerpAPI：請先安裝 google-search-results 套件")

    # Travelpayouts（可選，設定 TRAVELPAYOUTS_TOKEN 後啟用）
    tp_token = os.environ.get("TRAVELPAYOUTS_TOKEN", "").strip()
    if tp_token:
        from backends.travelpayouts_backend import TravelpayoutsBackend
        backends.append(TravelpayoutsBackend(tp_token))
        print("  ✓ Travelpayouts 已載入")

    if not backends:
        print("\n[錯誤] 沒有可用的後端。")
        print("  請安裝 fast-flights：pip install fast-flights")
        print("  或使用 Mock 模式測試：python flight_search.py --mock")
        sys.exit(1)

    return backends


# ── 搜尋與篩選 ─────────────────────────────────────────────────────────────────

def search_all(
    backends: list,
    origins: list[str],
    destinations: list[str],
    dest_name_map: dict,
    departure_date: str,
    return_date: str | None,
    currency: str,
    adults: int,
) -> list[FlightResult]:
    all_results: list[FlightResult] = []
    for backend in backends:
        label = type(backend).__name__.replace("Backend", "")
        for orig in origins:
            for dest in destinations:
                dest_name = dest_name_map.get(dest, dest)
                print(f"  [{label}] {orig} → {dest} ({dest_name})...", end=" ", flush=True)
                results = backend.search(orig, dest, dest_name, departure_date, return_date, currency, adults)
                print(f"{len(results)} 筆")
                all_results.extend(results)
    return all_results


def search_split(
    backends: list,
    origins: list[str],
    destinations: list[str],
    dest_name_map: dict,
    departure_date: str,
    return_date: str,
    currency: str,
    adults: int,
) -> tuple[list[FlightResult], list[FlightResult]]:
    """去回程分別以單程搜尋，取得各自票價"""
    print("  ▶ 去程搜尋")
    outbound = search_all(backends, origins, destinations, dest_name_map,
                          departure_date, None, currency, adults)
    print("  ▶ 回程搜尋")
    # 回程：目的地→出發地，dest_name_map 以代碼本身為名稱
    ret_name_map = {orig: orig for orig in origins}
    return_leg = search_all(backends, destinations, origins, ret_name_map,
                            return_date, None, currency, adults)
    return outbound, return_leg


def match_combinations(
    outbound: list[FlightResult],
    return_leg: list[FlightResult],
    currency: str,
) -> list[dict]:
    """以航空公司配對去回程，取各自最低價組合"""
    def cheapest_by_airline(lst: list[FlightResult]) -> dict:
        best: dict[str, FlightResult] = {}
        for r in lst:
            key = r.flights_str or r.source
            if key not in best or r.price < best[key].price:
                best[key] = r
        return best

    out_map = cheapest_by_airline(outbound)
    ret_map = cheapest_by_airline(return_leg)

    combos = []
    for airline, out in out_map.items():
        ret = ret_map.get(airline)
        if not ret:
            continue
        combos.append({
            "airline":      airline,
            "dep_time":     out.dep_time,
            "dep_arr":      out.arr_time,
            "dep_duration": out.duration,
            "dep_stops":    out.stops,
            "dep_price":    out.price,
            "ret_time":     ret.dep_time,
            "ret_arr":      ret.arr_time,
            "ret_duration": ret.duration,
            "ret_stops":    ret.stops,
            "ret_price":    ret.price,
            "total":        out.price + ret.price,
            "currency":     currency,
        })
    return sorted(combos, key=lambda x: x["total"])


def filter_by_budget(results: list[FlightResult], budget: float) -> list[FlightResult]:
    return [r for r in results if r.price <= budget]


# ── Terminal 輸出 ──────────────────────────────────────────────────────────────

def print_result(rank: int, r: FlightResult) -> None:
    stop_str = "直飛" if r.stops == 0 else ("不明" if r.stops < 0 else f"轉機 {r.stops} 次")
    print(f"  ┌─ #{rank}  {r.destination_name}  {r.price:,.0f} {r.currency}  [{stop_str}]  [{r.source}]")
    print(f"  │   去程出發: {r.dep_time}")
    if r.arr_time:
        print(f"  │   去程抵達: {r.arr_time}", end="")
        if r.duration:
            print(f"  (飛行 {r.duration})", end="")
        print()
    print(f"  │   航班: {r.flights_str}")
    if r.ret_dep_time:
        ret_stops_str = "" if r.ret_stops is None else f"  轉機 {r.ret_stops} 次" if r.ret_stops else "  直飛"
        print(f"  │   回程出發: {r.ret_dep_time}{ret_stops_str}")
    if r.note:
        print(f"  │   ※ {r.note}")
    print("  └─────────────────────────────────────────")


# ── Excel 樣式常數 ─────────────────────────────────────────────────────────────

_C_HEADER_BG  = "1F4E79"   # 深藍，標題列背景
_C_PARAM_BG   = "D6E4F0"   # 淺藍，搜尋條件背景
_C_DIRECT_BG  = "E2EFDA"   # 淺綠，直飛列
_C_STOP_BG    = "FFF2CC"   # 淺黃，轉機列
_C_BEST_BG    = "C6EFCE"   # 深綠，最低價前3名
_C_WHITE      = "FFFFFF"
_C_DEP_PRICE  = "BDD7EE"   # 藍，去程票價欄
_C_RET_PRICE  = "F9CBAD"   # 橘，回程票價欄
_C_TOT_PRICE  = "A9D18E"   # 綠，合計票價欄（一般）
_C_TOT_BEST   = "70AD47"   # 深綠，合計前3名

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


# ── Excel 輸出 ─────────────────────────────────────────────────────────────────

def _write_flights_sheet(
    ws,
    results: list[FlightResult],
    price_label: str,
    search_info: list[str],
) -> None:
    """單一 sheet 寫入：搜尋條件 + 標題列 + 資料列"""
    border = _thin_border()

    # 搜尋條件摘要（前幾行）
    param_font = Font(bold=True, size=10, color="1F4E79")
    param_fill = _fill(_C_PARAM_BG)
    for line in search_info:
        ws.append([line])
        ws.cell(ws.max_row, 1).font = param_font
        ws.cell(ws.max_row, 1).fill = param_fill
    ws.append([])

    # 標題列
    HEADER_ROW = ws.max_row + 1
    headers = ["排名", "航空公司", "出發時間", "抵達時間", "飛行時長", "直飛/轉機", price_label]
    ws.append(headers)
    hdr_font = Font(bold=True, size=11, color=_C_WHITE)
    hdr_fill = _fill(_C_HEADER_BG)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(HEADER_ROW, c)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = border
        cell.alignment = _center()
    ws.row_dimensions[HEADER_ROW].height = 22

    # 資料列
    top3_prices = sorted(set(r.price for r in results))[:3]
    lowest = results[0].price if results else 0

    for rank, r in enumerate(results, 1):
        stop_str = "直飛" if r.stops == 0 else ("不明" if r.stops < 0 else f"轉機 {r.stops} 次")
        row_data = [rank, r.flights_str or r.source, r.dep_time, r.arr_time,
                    r.duration, stop_str, r.price]
        ws.append(row_data)
        dr = ws.max_row
        ws.row_dimensions[dr].height = 18

        row_fill = (
            _fill(_C_BEST_BG)   if r.price in top3_prices else
            _fill(_C_DIRECT_BG) if r.stops == 0 else
            _fill(_C_STOP_BG)
        )
        for c in range(1, 8):
            cell = ws.cell(dr, c)
            cell.border = border
            cell.font   = Font(size=12)
            if c == 7:
                cell.fill = _fill(_C_DEP_PRICE)
                cell.alignment = _right()
                cell.number_format = "#,##0"
                if r.price == lowest:
                    cell.font = Font(bold=True, size=13, color="C00000")
            else:
                cell.fill = row_fill
                cell.alignment = _center() if c == 1 else Alignment(vertical="center")

    # 欄寬
    for col_idx, w in {1:6, 2:20, 3:20, 4:20, 5:12, 6:12, 7:16}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 凍結 + 篩選
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:G{ws.max_row}"

    # 色碼圖例
    lr = ws.max_row + 2
    for i, (color, label) in enumerate([(_C_BEST_BG, "最低價前3名"),
                                         (_C_DIRECT_BG, "直飛"),
                                         (_C_STOP_BG, "含轉機")]):
        c = i * 2 + 1
        ws.cell(lr, c).fill   = _fill(color)
        ws.cell(lr, c).border = _thin_border()
        ws.cell(lr, c + 1).value = label
        ws.cell(lr, c + 1).font  = Font(size=9, color="595959")


def _write_combo_sheet(ws, combos: list[dict], search_params: dict) -> None:
    """組合推薦 sheet：去程 + 回程 + 合計"""
    border = _thin_border()
    currency = combos[0]["currency"] if combos else "TWD"

    # 標題
    ws.append([f"✈ 組合推薦（同航空公司去回程最低價）  共 {len(combos)} 組"])
    ws["A1"].font = Font(bold=True, size=13, color=_C_HEADER_BG)
    ws.merge_cells("A1:I1")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 24
    ws.append([])

    # 標題列
    HEADER_ROW = 3
    headers = [
        "排名", "航空公司",
        "去程出發", "去程抵達", "去程時長", "去程轉機", f"去程票價({currency})",
        f"回程出發", "回程抵達", "回程時長", "回程轉機", f"回程票價({currency})",
        f"合計票價({currency})",
    ]
    ws.append(headers)
    hdr_font = Font(bold=True, size=11, color=_C_WHITE)
    hdr_fill = _fill(_C_HEADER_BG)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(HEADER_ROW, c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.border = border; cell.alignment = _center(wrap=True)
    ws.row_dimensions[HEADER_ROW].height = 28

    # 資料列
    top3 = sorted(set(c["total"] for c in combos))[:3]
    lowest = combos[0]["total"] if combos else 0

    def stop_str(stops):
        return "直飛" if stops == 0 else ("不明" if stops < 0 else f"轉機{stops}次")

    for rank, combo in enumerate(combos, 1):
        row = [
            rank, combo["airline"],
            combo["dep_time"], combo["dep_arr"], combo["dep_duration"],
            stop_str(combo["dep_stops"]), combo["dep_price"],
            combo["ret_time"], combo["ret_arr"], combo["ret_duration"],
            stop_str(combo["ret_stops"]), combo["ret_price"],
            combo["total"],
        ]
        ws.append(row)
        dr = ws.max_row
        ws.row_dimensions[dr].height = 18

        row_fill = _fill(_C_BEST_BG) if combo["total"] in top3 else _fill(_C_DIRECT_BG)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(dr, c)
            cell.fill = row_fill; cell.border = border; cell.font = Font(size=10)
            if c == 1:
                cell.alignment = _center()
            elif c in (7, 12, 13):
                cell.alignment = _right()
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(vertical="center")
        # 合計欄加粗
        ws.cell(dr, 13).font = Font(bold=True, size=11,
                                     color="C00000" if combo["total"] == lowest else "1F4E79")

    # 欄寬
    col_w = {1:6, 2:20, 3:20, 4:20, 5:12, 6:10, 7:16, 8:20, 9:20, 10:12, 11:10, 12:16, 13:18}
    for ci, w in col_w.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{ws.max_row}"


def export_to_excel(
    search_params: dict,
    results: list[FlightResult],
    return_leg: list[FlightResult] | None = None,
    combinations: list[dict] | None = None,
) -> str:
    output_dir = os.path.join(SCRIPT_DIR, "output")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filepath  = os.path.join(output_dir, f"flights_{timestamp}.xlsx")

    currency  = results[0].currency if results else "TWD"
    origin    = search_params.get("出發機場", "")
    dest      = search_params.get("目的地", "")
    dep_date  = search_params.get("去程日期", "")
    ret_date  = search_params.get("回程日期", "")
    pax       = search_params.get("乘客人數", "1")
    budget    = search_params.get("預算上限", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    border = _thin_border()

    if return_leg is not None and combinations is not None:
        # ── 去回程：單頁，含去程/回程/合計三欄票價 ──────
        ws.title = "航班搜尋結果"

        # 搜尋條件
        param_font = Font(bold=True, size=10, color="1F4E79")
        param_fill = _fill(_C_PARAM_BG)
        info_lines = [
            f"✈  {origin} → {dest}   去程: {dep_date}   回程: {ret_date}   乘客: {pax} 人   預算: {budget} {currency}/人",
            f"組合推薦 {len(combinations)} 組（同航空公司去回程最低價）  |  資料來源: Google Flights  |  票價僅供參考，請至官網確認",
        ]
        for line in info_lines:
            ws.append([line])
            r = ws.max_row
            ws.cell(r, 1).font = param_font
            ws.cell(r, 1).fill = param_fill
        ws.append([])

        # 標題列
        HEADER_ROW = 4
        headers = [
            "排名", "航空公司",
            "去程出發", "去程抵達", "去程時長", "去程轉機", f"去程票價({currency})",
            "回程出發", "回程抵達", "回程時長", "回程轉機", f"回程票價({currency})",
            f"合計票價({currency})",
        ]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(HEADER_ROW, c)
            cell.font = Font(bold=True, size=11, color=_C_WHITE)
            cell.fill = _fill(_C_HEADER_BG)
            cell.border = border
            cell.alignment = _center(wrap=True)
        ws.row_dimensions[HEADER_ROW].height = 30

        # 資料列
        top3 = sorted(set(c["total"] for c in combinations))[:3]
        lowest = combinations[0]["total"] if combinations else 0

        def _stop(s): return "直飛" if s == 0 else ("不明" if s < 0 else f"轉機{s}次")

        for rank, c in enumerate(combinations, 1):
            row = [
                rank, c["airline"],
                c["dep_time"], c["dep_arr"], c["dep_duration"], _stop(c["dep_stops"]), c["dep_price"],
                c["ret_time"], c["ret_arr"], c["ret_duration"], _stop(c["ret_stops"]), c["ret_price"],
                c["total"],
            ]
            ws.append(row)
            dr = ws.max_row
            ws.row_dimensions[dr].height = 20

            is_top3 = c["total"] in top3
            row_fill = _fill(_C_BEST_BG) if is_top3 else _fill(_C_DIRECT_BG)

            for ci in range(1, len(headers) + 1):
                cell = ws.cell(dr, ci)
                cell.border = border
                cell.font   = Font(size=12)

                # 票價欄各自獨立底色
                if ci == 7:
                    cell.fill = _fill(_C_DEP_PRICE)
                    cell.alignment = _right(); cell.number_format = "#,##0"
                elif ci == 12:
                    cell.fill = _fill(_C_RET_PRICE)
                    cell.alignment = _right(); cell.number_format = "#,##0"
                elif ci == 13:
                    cell.fill = _fill(_C_TOT_BEST if is_top3 else _C_TOT_PRICE)
                    cell.alignment = _right(); cell.number_format = "#,##0"
                    cell.font = Font(bold=True, size=13,
                                     color="C00000" if c["total"] == lowest else "375623")
                else:
                    cell.fill = row_fill
                    cell.alignment = _center() if ci == 1 else Alignment(vertical="center")

        # 欄寬
        for ci, w in {1:6, 2:20, 3:20, 4:20, 5:12, 6:10, 7:16,
                      8:20, 9:20, 10:12, 11:10, 12:16, 13:18}.items():
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = f"A{HEADER_ROW + 1}"
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{ws.max_row}"

        # 色碼圖例
        lr = ws.max_row + 2
        for i, (color, label) in enumerate([(_C_BEST_BG, "合計最低前3名"), (_C_DIRECT_BG, "其餘組合")]):
            c_idx = i * 2 + 1
            ws.cell(lr, c_idx).fill = _fill(color); ws.cell(lr, c_idx).border = _thin_border()
            ws.cell(lr, c_idx + 1).value = label
            ws.cell(lr, c_idx + 1).font = Font(size=9, color="595959")

    else:
        # ── 單程：單頁 ────────────────────────────────────
        ws.title = "航班搜尋結果"
        info = [
            f"✈ 單程  {origin} → {dest}   出發: {dep_date}   乘客: {pax} 人   預算: {budget} {currency}/人",
            f"共 {len(results)} 筆  |  資料來源: Google Flights  |  票價僅供參考，請至官網確認",
        ]
        _write_flights_sheet(ws, results, f"票價({currency})", info)

    wb.save(filepath)
    return filepath


# ── 使用者輸入工具 ─────────────────────────────────────────────────────────────

def choose_from_menu(title: str, options: dict, allow_all: bool = False) -> list[str]:
    print(f"\n{title}")
    for key, (name, code) in options.items():
        print(f"  {key:>2}. {name} ({code})")
    if allow_all:
        print("   0. 全部搜尋")

    while True:
        raw = input("請輸入編號（多個用逗號分隔，例 1,3）: ").strip()
        if allow_all and raw == "0":
            return [code for _, code in options.values()]
        keys  = [k.strip() for k in raw.split(",")]
        codes = [options[k][1] for k in keys if k in options]
        if codes:
            return codes
        print("  輸入無效，請重試。")


def ask_date(prompt: str) -> str:
    while True:
        raw = input(f"{prompt}: ").strip()
        try:
            d = datetime.strptime(raw, "%Y-%m-%d").date()
            if d < date.today():
                print("  日期不能早於今天，請重試。")
                continue
            return raw
        except ValueError:
            print("  格式錯誤，請輸入如 2026-08-01。")


def ask_return_date(departure_date: str) -> str | None:
    dep = datetime.strptime(departure_date, "%Y-%m-%d").date()
    while True:
        raw = input("回程日期（YYYY-MM-DD，Enter 略過=單程）: ").strip()
        if not raw:
            return None
        try:
            ret = datetime.strptime(raw, "%Y-%m-%d").date()
            if ret <= dep:
                print("  回程日期必須晚於去程日期，請重試。")
                continue
            return raw
        except ValueError:
            print("  格式錯誤，請輸入如 2026-08-10。")


def ask_budget(currency_code: str) -> float:
    while True:
        raw = input(f"預算上限（{currency_code}，每人）: ").strip()
        try:
            val = float(raw.replace(",", ""))
            if val > 0:
                return val
        except ValueError:
            pass
        print("  請輸入有效的數字。")


def ask_int(prompt: str, default: int) -> int:
    raw = input(f"{prompt}（預設 {default}）: ").strip()
    try:
        val = int(raw)
        return val if val > 0 else default
    except ValueError:
        return default


# ── 主程式 ────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="日本航班搜尋工具")
    parser.add_argument("--mock", action="store_true", help="使用假資料測試，不需要任何 API Key")
    args = parser.parse_args()

    print("=" * 52)
    print("   日本航班搜尋工具  (Google Flights)")
    print("=" * 52)

    print("\n【載入 Backend】")
    backends = build_backends(mock_mode=args.mock)

    # 1. 出發機場
    origins = choose_from_menu("【出發機場】", ORIGIN_AIRPORTS)

    # 2. 目的地
    destinations = choose_from_menu("【目的地（日本）】", JAPAN_AIRPORTS, allow_all=True)
    dest_name_map = {code: name for _, (name, code) in JAPAN_AIRPORTS.items()}

    # 3. 去回程日期
    print()
    departure_date = ask_date("去程日期（YYYY-MM-DD）")
    return_date    = ask_return_date(departure_date)

    # 4. 幣別
    currency_codes = choose_from_menu("【幣別】", CURRENCY_OPTIONS)
    currency = currency_codes[0]

    # 5. 預算 & 人數
    print()
    budget = ask_budget(currency)
    adults = ask_int("乘客人數", 1)

    trip_type = "去回程（分段搜尋）" if return_date else "單程"
    print(f"\n正在搜尋 {departure_date} {trip_type}，幣別 {currency}，預算 {budget:,.0f}，{adults} 位乘客...\n")

    search_params = {
        "出發機場": ", ".join(origins),
        "目的地":   ", ".join(dest_name_map.get(d, d) for d in destinations),
        "去程日期": departure_date,
        "回程日期": return_date or "（單程）",
        "幣別":     currency,
        "預算上限": f"{budget:,.0f}",
        "乘客人數": str(adults),
    }

    if return_date:
        # 去回程：分別單程搜尋
        raw_out, raw_ret = search_split(
            backends, origins, destinations, dest_name_map,
            departure_date, return_date, currency, adults,
        )
        aff_out = sorted(filter_by_budget(raw_out, budget), key=lambda r: r.price)
        aff_ret = sorted(filter_by_budget(raw_ret, budget), key=lambda r: r.price)
        combos  = match_combinations(aff_out, aff_ret, currency)

        if not aff_out and not aff_ret:
            print(f"\n找不到預算 {budget:,.0f} {currency} 以內的航班。")
            return

        print(f"\n{'=' * 52}")
        print(f"  去程：共 {len(aff_out)} 筆  |  回程：共 {len(aff_ret)} 筆")
        print(f"  組合推薦：{len(combos)} 組（同航空公司最低價）")
        print(f"{'=' * 52}\n")

        if combos:
            print("  【組合推薦 Top 5】")
            for i, c in enumerate(combos[:5], 1):
                print(f"  #{i}  {c['airline']:20s}  去程 {c['dep_price']:>8,.0f}  "
                      f"+ 回程 {c['ret_price']:>8,.0f}  = {c['total']:>9,.0f} {currency}")
        if aff_out:
            print(f"\n  【去程最低】 {aff_out[0].flights_str}  {aff_out[0].price:,.0f} {currency}")
        if aff_ret:
            print(f"  【回程最低】 {aff_ret[0].flights_str}  {aff_ret[0].price:,.0f} {currency}")

        search_params["符合預算筆數"] = f"去程 {len(aff_out)} / 回程 {len(aff_ret)}"
        output_path = export_to_excel(search_params, aff_out, aff_ret, combos)

    else:
        # 單程
        raw = search_all(backends, origins, destinations, dest_name_map,
                         departure_date, None, currency, adults)
        affordable = sorted(filter_by_budget(raw, budget), key=lambda r: r.price)

        if not affordable:
            print(f"\n找不到預算 {budget:,.0f} {currency} 以內的航班。")
            return

        print(f"\n{'=' * 52}")
        print(f"  共找到 {len(affordable)} 個符合預算的航班（顯示前 10 筆）")
        print(f"{'=' * 52}\n")
        for rank, r in enumerate(affordable[:10], 1):
            print_result(rank, r)
        print(f"\n最低: {affordable[0].price:,.0f} {currency}  → {affordable[0].destination_name}")

        search_params["符合預算筆數"] = str(len(affordable))
        output_path = export_to_excel(search_params, affordable)

    print(f"\nExcel 已輸出：{output_path}")


if __name__ == "__main__":
    main()
