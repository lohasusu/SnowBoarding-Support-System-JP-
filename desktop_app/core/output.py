"""Excel output writers (openpyxl-based, no Playwright/heavy deps)."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _header(ws, row: int, cols: list[str]) -> None:
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = f"A{row + 1}"


def _autofit(ws, max_widths: dict[int, int] | None = None) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        m = max((len(str(c.value or "")) for c in col), default=8)
        width = min(max(m + 2, 8), 50)
        if max_widths and col[0].column in max_widths:
            width = max_widths[col[0].column]
        ws.column_dimensions[letter].width = width


def write_ski(items: list[dict], out_dir: Path, params: dict) -> Path:
    """Write ski tickets → ski_{ts}.xlsx; return file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir.mkdir(parents=True, exist_ok=True)
    fp = out_dir / f"ski_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "雪票"

    # Meta row
    info = (f"⛷  查詢條件: region={params.get('region') or '全部'} "
            f"name={params.get('name') or '無'}  |  "
            f"產生於 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  共 {len(items)} 筆")
    ws.append([info])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1).fill = PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid")
    ws.cell(row=1, column=1).font = Font(bold=True, size=10, color="1F4E79")

    cols = ["雪場", "地區", "票種（日文）", "票種（中文）", "票價", "雪季", "票價頁 URL"]
    _header(ws, row=3, cols=cols)

    for r in items:
        ws.append([
            r.get("resort", ""),
            r.get("region", ""),
            r.get("ticket_type", ""),
            r.get("ticket_type_zh", ""),
            r.get("price", ""),
            r.get("season", ""),
            r.get("source_url", ""),
        ])
    _autofit(ws, max_widths={7: 60})
    wb.save(fp)
    return fp


def write_flight(items: list[dict], out_dir: Path, params: dict) -> Path:
    """Write flight results → flight_{ts}.xlsx; return file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir.mkdir(parents=True, exist_ok=True)
    fp = out_dir / f"flight_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "機票"

    info = (f"✈  {params.get('origin','')} → {params.get('destination','')} "
            f"({params.get('dest_name','')})   "
            f"去程 {params.get('departure','')}"
            + (f"  回程 {params.get('ret_date')}" if params.get("ret_date") else "")
            + f"   乘客 {params.get('adults', 1)} 人  |  共 {len(items)} 筆  |  "
            f"產生於 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.append([info])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).fill = PatternFill(start_color="E7F4EE", end_color="E7F4EE", fill_type="solid")
    ws.cell(row=1, column=1).font = Font(bold=True, size=10, color="1B6E3F")

    cols = ["航空公司", "航班", "去程出發", "去程抵達", "飛行時間", "轉機", "票價 (TWD)", "資料來源"]
    _header(ws, row=3, cols=cols)

    for r in items:
        airlines = r.get("airline_names") or []
        airline = " / ".join(set(airlines)) if airlines else "未知"
        ws.append([
            airline,
            r.get("flights_str", ""),
            r.get("dep_time", ""),
            r.get("arr_time", ""),
            r.get("duration", ""),
            f"{r.get('stops', 0)} 轉" if (r.get("stops") or 0) > 0 else "直飛",
            r.get("price", ""),
            r.get("source", ""),
        ])
    _autofit(ws, max_widths={2: 50})
    wb.save(fp)
    return fp
